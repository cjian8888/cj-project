#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
电子钱包专项产物构建器

生成三类独立产物：
1. 电子钱包补充分析报告.txt
2. 电子钱包补充清洗表.xlsx
3. 电子钱包重点核查清单.txt
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Any, Dict, Iterable, List

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

import utils
from utils.safe_types import safe_float, safe_int, safe_str

logger = utils.setup_logger(__name__)

WALLET_TXT_REPORT_FILE = "电子钱包补充分析报告.txt"
WALLET_EXCEL_REPORT_FILE = "电子钱包补充清洗表.xlsx"
WALLET_FOCUS_REPORT_FILE = "电子钱包重点核查清单.txt"

_HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
_HEADER_FONT = Font(bold=True)
_WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)


def generate_wallet_artifacts(
    output_dir: str,
    wallet_data: Dict[str, Any],
    artifact_details: Dict[str, List[Dict[str, Any]]],
) -> Dict[str, str]:
    """生成电子钱包专项 TXT、Excel 与轻量核查清单。"""
    if not isinstance(wallet_data, dict) or not wallet_data.get("available"):
        logger.info("电子钱包专项产物跳过: 未检测到可用电子钱包数据")
        return {}

    os.makedirs(output_dir, exist_ok=True)
    txt_path = os.path.join(output_dir, WALLET_TXT_REPORT_FILE)
    excel_path = os.path.join(output_dir, WALLET_EXCEL_REPORT_FILE)
    focus_txt_path = os.path.join(output_dir, WALLET_FOCUS_REPORT_FILE)

    generate_wallet_txt_report(txt_path, wallet_data, artifact_details)
    generate_wallet_excel_workbook(excel_path, wallet_data, artifact_details)
    generate_wallet_focus_txt_report(focus_txt_path, wallet_data)
    return {"txt": txt_path, "excel": excel_path, "focus_txt": focus_txt_path}


def generate_wallet_txt_report(
    output_path: str,
    wallet_data: Dict[str, Any],
    artifact_details: Dict[str, List[Dict[str, Any]]],
) -> str:
    """生成电子钱包专项 TXT 分析报告。"""
    summary = wallet_data.get("summary", {}) or {}
    source_stats = wallet_data.get("sourceStats", {}) or {}
    source_files = artifact_details.get("sourceFiles", []) or []
    subjects = wallet_data.get("subjects", []) or []
    alerts = wallet_data.get("alerts", []) or []
    unmatched = wallet_data.get("unmatchedWechatAccounts", []) or []

    lines: List[str] = []
    lines.append("电子钱包补充分析报告")
    lines.append("=" * 40)
    lines.append("")
    lines.append("一、数据定位")
    lines.append("本报告对应微信 / 支付宝 / 财付通补充样本，定位为主链银行流水之外的后补证据层。")
    lines.append("电子钱包数据不写入 cleaned_data，只用于增强画像、预警、图谱和专项核查。")
    lines.append("")
    lines.append("二、样本文件统计")
    for label, value in (
        ("支付宝注册文件数", source_stats.get("alipayRegistrationFiles", 0)),
        ("支付宝账户明细文件数", source_stats.get("alipayTransactionFiles", 0)),
        ("微信注册文件数", source_stats.get("wechatRegistrationFiles", 0)),
        ("微信登录轨迹文件数", source_stats.get("wechatLoginFiles", 0)),
        ("财付通注册文件数", source_stats.get("tenpayRegistrationFiles", 0)),
        ("财付通交易文件数", source_stats.get("tenpayTransactionFiles", 0)),
        ("来源文件总数", len(source_files)),
    ):
        lines.append(f"- {label}: {value}")

    if source_files:
        lines.append("")
        lines.append("三、来源文件清单")
        for item in source_files:
            data_type = safe_str(item.get("dataType")) or "未知类型"
            relative_path = safe_str(item.get("relativePath")) or safe_str(item.get("fileName")) or ""
            lines.append(f"- [{data_type}] {relative_path}")

    lines.append("")
    lines.append("四、总体样本指标")
    for label, value in (
        ("识别主体数", summary.get("subjectCount", 0)),
        ("命中主链主体数", summary.get("coreMatchedSubjectCount", 0)),
        ("支付宝实名账户数", summary.get("alipayAccountCount", 0)),
        ("支付宝有效交易笔数", summary.get("alipayTransactionCount", 0)),
        ("微信账号数", summary.get("wechatAccountCount", 0)),
        ("财付通账号数", summary.get("tenpayAccountCount", 0)),
        ("财付通交易笔数", summary.get("tenpayTransactionCount", 0)),
        ("微信登录事件数", summary.get("loginEventCount", 0)),
        ("未归并微信账号数", summary.get("unmatchedWechatCount", 0)),
    ):
        lines.append(f"- {label}: {value}")

    lines.append("")
    lines.append("五、主体摘要")
    if subjects:
        for index, subject in enumerate(subjects[:10], start=1):
            subject_name = safe_str(subject.get("subjectName")) or safe_str(subject.get("subjectId")) or "未知主体"
            cross = subject.get("crossSignals", {}) or {}
            platforms = subject.get("platforms", {}) or {}
            alipay = platforms.get("alipay", {}) or {}
            wechat = platforms.get("wechat", {}) or {}
            total_count = _int_value(alipay.get("transactionCount")) + _int_value(
                wechat.get("tenpayTransactionCount")
            )
            total_amount = _float_value(alipay.get("incomeTotalYuan")) + _float_value(
                alipay.get("expenseTotalYuan")
            ) + _float_value(wechat.get("incomeTotalYuan")) + _float_value(
                wechat.get("expenseTotalYuan")
            )
            lines.append(
                f"{index}. {subject_name} | 证件号: {safe_str(subject.get('subjectId')) or '-'} | "
                f"主链命中: {'是' if subject.get('matchedToCore') else '否'} | "
                f"总交易笔数: {total_count} | 总规模: {total_amount:.2f} 元"
            )
            lines.append(
                "   支付宝: "
                f"{_int_value(alipay.get('accountCount'))} 账户 / "
                f"{_int_value(alipay.get('transactionCount'))} 笔 / "
                f"收入 {_float_value(alipay.get('incomeTotalYuan')):.2f} / "
                f"支出 {_float_value(alipay.get('expenseTotalYuan')):.2f}"
            )
            lines.append(
                "   微信/财付通: "
                f"{_int_value(wechat.get('wechatAccountCount'))} 微信号 / "
                f"{_int_value(wechat.get('tenpayAccountCount'))} 财付通账号 / "
                f"{_int_value(wechat.get('tenpayTransactionCount'))} 笔 / "
                f"登录 {_int_value(wechat.get('loginEventCount'))} 次"
            )
            lines.append(
                "   跨平台信号: "
                f"手机号重叠 {_int_value(cross.get('phoneOverlapCount'))} 组, "
                f"银行卡重叠 {_int_value(cross.get('bankCardOverlapCount'))} 张, "
                f"别名重叠 {_int_value(cross.get('aliasMatchCount'))} 组"
            )
    else:
        lines.append("- 未识别到主体摘要。")

    lines.append("")
    lines.append("六、重点电子钱包预警")
    if alerts:
        for index, alert in enumerate(alerts[:20], start=1):
            lines.append(
                f"{index}. [{safe_str(alert.get('risk_level')) or 'medium'}] "
                f"{safe_str(alert.get('person')) or '未知主体'} -> "
                f"{safe_str(alert.get('counterparty')) or '未知对手方'} | "
                f"金额 {_float_value(alert.get('amount')):.2f} 元 | "
                f"{safe_str(alert.get('description')) or ''}"
            )
    else:
        lines.append("- 暂无专项预警。")

    lines.append("")
    lines.append("七、未归并微信账号")
    if unmatched:
        for item in unmatched:
            lines.append(
                f"- 手机号 {safe_str(item.get('phone')) or '-'} | "
                f"微信号 {safe_str(item.get('wxid')) or '-'} | "
                f"别名 {safe_str(item.get('alias')) or '-'} | "
                f"昵称 {safe_str(item.get('nickname')) or '-'} | "
                f"最近登录 {safe_str(item.get('latestLoginAt')) or '-'} | "
                f"登录次数 {_int_value(item.get('loginEventCount'))}"
            )
    else:
        lines.append("- 无未归并微信账号。")

    lines.append("")
    lines.append("八、建议用法")
    for suggestion in _build_wallet_suggestions(wallet_data):
        lines.append(f"- {suggestion}")

    Path(output_path).write_text("\n".join(lines), encoding="utf-8")
    return output_path


def generate_wallet_excel_workbook(
    output_path: str,
    wallet_data: Dict[str, Any],
    artifact_details: Dict[str, List[Dict[str, Any]]],
) -> str:
    """生成电子钱包专项 Excel 清洗表。"""
    workbook_data = {
        "样本概览": _build_overview_df(wallet_data, artifact_details),
        "来源文件清单": _build_source_file_df(artifact_details),
        "主体汇总": _build_subject_summary_df(wallet_data),
        "支付宝实名账户": _build_alipay_registration_df(wallet_data, artifact_details),
        "支付宝交易明细": _build_alipay_transaction_df(wallet_data, artifact_details),
        "微信注册信息": _build_wechat_registration_df(wallet_data, artifact_details),
        "微信登录轨迹": _build_wechat_login_df(wallet_data, artifact_details),
        "财付通实名账户": _build_tenpay_registration_df(wallet_data, artifact_details),
        "财付通交易明细": _build_tenpay_transaction_df(wallet_data, artifact_details),
        "电子钱包预警": _build_alert_df(wallet_data),
        "未归并微信账号": _build_unmatched_df(wallet_data),
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in workbook_data.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        for sheet_name, df in workbook_data.items():
            worksheet = workbook[sheet_name]
            _format_worksheet(worksheet, len(df.columns))

    return output_path


def generate_wallet_focus_txt_report(
    output_path: str,
    wallet_data: Dict[str, Any],
) -> str:
    """生成轻量电子钱包重点核查清单。"""
    summary = wallet_data.get("summary", {}) or {}
    unmatched = wallet_data.get("unmatchedWechatAccounts", []) or []
    high_risk_alerts = [
        item
        for item in (wallet_data.get("alerts", []) or [])
        if (safe_str(item.get("risk_level")) or "").lower() == "high"
    ]

    lines: List[str] = []
    lines.append("电子钱包重点核查清单")
    lines.append("=" * 32)
    lines.append("")
    lines.append("一、定位")
    lines.append("本清单为轻量输出，仅保留高风险电子钱包预警和未归并微信账号，便于快速定位下一步人工核查重点。")
    lines.append("")
    lines.append("二、速览")
    lines.append(f"- 高风险电子钱包预警数: {len(high_risk_alerts)}")
    lines.append(f"- 未归并微信账号数: {_int_value(summary.get('unmatchedWechatCount'))}")
    lines.append("")
    lines.append("三、高风险电子钱包预警")
    if high_risk_alerts:
        for index, alert in enumerate(high_risk_alerts, start=1):
            lines.append(
                f"{index}. {safe_str(alert.get('person')) or '未知主体'} -> "
                f"{safe_str(alert.get('counterparty')) or '未知对手方'} | "
                f"日期 {safe_str(alert.get('date')) or '-'} | "
                f"金额 {_float_value(alert.get('amount')):.2f} 元"
            )
            lines.append(f"   描述: {safe_str(alert.get('description')) or '-'}")
            lines.append(f"   原因: {safe_str(alert.get('risk_reason')) or '-'}")
    else:
        lines.append("- 当前未发现 high 级别电子钱包预警。")

    lines.append("")
    lines.append("四、未归并微信账号")
    if unmatched:
        for index, item in enumerate(unmatched, start=1):
            lines.append(
                f"{index}. 手机号 {safe_str(item.get('phone')) or '-'} | "
                f"微信号 {safe_str(item.get('wxid')) or '-'} | "
                f"别名 {safe_str(item.get('alias')) or '-'} | "
                f"昵称 {safe_str(item.get('nickname')) or '-'}"
            )
            lines.append(
                f"   注册时间 {safe_str(item.get('registeredAt')) or '-'} | "
                f"最近登录 {safe_str(item.get('latestLoginAt')) or '-'} | "
                f"登录次数 {_int_value(item.get('loginEventCount'))}"
            )
    else:
        lines.append("- 无未归并微信账号。")

    Path(output_path).write_text("\n".join(lines), encoding="utf-8")
    return output_path


def _build_wallet_suggestions(wallet_data: Dict[str, Any]) -> List[str]:
    summary = wallet_data.get("summary", {}) or {}
    alerts = wallet_data.get("alerts", []) or []
    suggestions = [
        "优先将命中主链主体的电子钱包对手方，与银行流水高频对手方做交叉核验。",
        "支付宝与财付通的跨平台手机号、银行卡、别名重叠，可作为同一控制人的补强证据。",
        "对未归并微信账号，建议补充手机号、别名、设备或实名映射后再次导入分析。",
    ]
    if _int_value(summary.get("coreMatchedSubjectCount")) == 0:
        suggestions.append("当前样本尚未与主链主体形成稳定映射，应优先补齐姓名或证件号映射。")
    if any((safe_str(item.get("risk_level")) or "").lower() == "high" for item in alerts):
        suggestions.append("高风险电子钱包预警已出现，建议把相关主体纳入下一轮明细核查清单。")
    return suggestions


def _build_overview_df(
    wallet_data: Dict[str, Any],
    artifact_details: Dict[str, List[Dict[str, Any]]],
) -> pd.DataFrame:
    source_stats = wallet_data.get("sourceStats", {}) or {}
    summary = wallet_data.get("summary", {}) or {}
    notes = wallet_data.get("notes", []) or []
    rows = [
        {"类别": "文件统计", "指标": "支付宝注册文件数", "值": source_stats.get("alipayRegistrationFiles", 0)},
        {"类别": "文件统计", "指标": "支付宝账户明细文件数", "值": source_stats.get("alipayTransactionFiles", 0)},
        {"类别": "文件统计", "指标": "微信注册文件数", "值": source_stats.get("wechatRegistrationFiles", 0)},
        {"类别": "文件统计", "指标": "微信登录轨迹文件数", "值": source_stats.get("wechatLoginFiles", 0)},
        {"类别": "文件统计", "指标": "财付通注册文件数", "值": source_stats.get("tenpayRegistrationFiles", 0)},
        {"类别": "文件统计", "指标": "财付通交易文件数", "值": source_stats.get("tenpayTransactionFiles", 0)},
        {"类别": "文件统计", "指标": "来源文件总数", "值": len(artifact_details.get("sourceFiles", []) or [])},
        {"类别": "总体指标", "指标": "识别主体数", "值": summary.get("subjectCount", 0)},
        {"类别": "总体指标", "指标": "命中主链主体数", "值": summary.get("coreMatchedSubjectCount", 0)},
        {"类别": "总体指标", "指标": "支付宝实名账户数", "值": summary.get("alipayAccountCount", 0)},
        {"类别": "总体指标", "指标": "支付宝有效交易笔数", "值": summary.get("alipayTransactionCount", 0)},
        {"类别": "总体指标", "指标": "微信账号数", "值": summary.get("wechatAccountCount", 0)},
        {"类别": "总体指标", "指标": "财付通账号数", "值": summary.get("tenpayAccountCount", 0)},
        {"类别": "总体指标", "指标": "财付通交易笔数", "值": summary.get("tenpayTransactionCount", 0)},
        {"类别": "总体指标", "指标": "微信登录事件数", "值": summary.get("loginEventCount", 0)},
        {"类别": "总体指标", "指标": "未归并微信账号数", "值": summary.get("unmatchedWechatCount", 0)},
    ]
    for index, note in enumerate(notes, start=1):
        rows.append({"类别": "说明", "指标": f"说明{index}", "值": note})
    return pd.DataFrame(rows, columns=["类别", "指标", "值"])


def _build_source_file_df(artifact_details: Dict[str, List[Dict[str, Any]]]) -> pd.DataFrame:
    rows = []
    for index, item in enumerate(artifact_details.get("sourceFiles", []) or [], start=1):
        rows.append(
            {
                "序号": index,
                "数据类型": safe_str(item.get("dataType")) or "",
                "相对路径": safe_str(item.get("relativePath")) or "",
                "文件名": safe_str(item.get("fileName")) or "",
                "原始路径": safe_str(item.get("filePath")) or "",
            }
        )
    return _frame(rows, ["序号", "数据类型", "相对路径", "文件名", "原始路径"])


def _build_subject_summary_df(wallet_data: Dict[str, Any]) -> pd.DataFrame:
    rows = []
    for index, subject in enumerate(wallet_data.get("subjects", []) or [], start=1):
        platforms = subject.get("platforms", {}) or {}
        alipay = platforms.get("alipay", {}) or {}
        wechat = platforms.get("wechat", {}) or {}
        cross = subject.get("crossSignals", {}) or {}
        total_count = _int_value(alipay.get("transactionCount")) + _int_value(
            wechat.get("tenpayTransactionCount")
        )
        total_amount = _float_value(alipay.get("incomeTotalYuan")) + _float_value(
            alipay.get("expenseTotalYuan")
        ) + _float_value(wechat.get("incomeTotalYuan")) + _float_value(
            wechat.get("expenseTotalYuan")
        )
        rows.append(
            {
                "序号": index,
                "主体姓名": safe_str(subject.get("subjectName")) or "",
                "主体证件号": safe_str(subject.get("subjectId")) or "",
                "是否命中主链": "是" if subject.get("matchedToCore") else "否",
                "总交易笔数": total_count,
                "总交易规模(元)": round(total_amount, 2),
                "支付宝账户数": _int_value(alipay.get("accountCount")),
                "支付宝交易笔数": _int_value(alipay.get("transactionCount")),
                "支付宝收入(元)": _float_value(alipay.get("incomeTotalYuan")),
                "支付宝支出(元)": _float_value(alipay.get("expenseTotalYuan")),
                "微信账号数": _int_value(wechat.get("wechatAccountCount")),
                "财付通账号数": _int_value(wechat.get("tenpayAccountCount")),
                "财付通交易笔数": _int_value(wechat.get("tenpayTransactionCount")),
                "财付通收入(元)": _float_value(wechat.get("incomeTotalYuan")),
                "财付通支出(元)": _float_value(wechat.get("expenseTotalYuan")),
                "登录事件数": _int_value(wechat.get("loginEventCount")),
                "手机号重叠组数": _int_value(cross.get("phoneOverlapCount")),
                "银行卡重叠张数": _int_value(cross.get("bankCardOverlapCount")),
                "别名重叠组数": _int_value(cross.get("aliasMatchCount")),
                "匹配依据": "；".join(subject.get("crossSignals", {}).get("matchBasis", []) or []),
                "补充信号": "；".join(subject.get("signals", []) or []),
            }
        )
    return _frame(
        rows,
        [
            "序号",
            "主体姓名",
            "主体证件号",
            "是否命中主链",
            "总交易笔数",
            "总交易规模(元)",
            "支付宝账户数",
            "支付宝交易笔数",
            "支付宝收入(元)",
            "支付宝支出(元)",
            "微信账号数",
            "财付通账号数",
            "财付通交易笔数",
            "财付通收入(元)",
            "财付通支出(元)",
            "登录事件数",
            "手机号重叠组数",
            "银行卡重叠张数",
            "别名重叠组数",
            "匹配依据",
            "补充信号",
        ],
    )


def _build_alipay_registration_df(
    wallet_data: Dict[str, Any],
    artifact_details: Dict[str, List[Dict[str, Any]]],
) -> pd.DataFrame:
    rows = []
    for index, row in enumerate(_enrich_subject_rows(wallet_data, artifact_details.get("alipayRegistrationRows", [])), start=1):
        rows.append(
            {
                "序号": index,
                "主体姓名": row["subjectName"],
                "主体证件号": row["subjectId"],
                "是否命中主链": _yes_no(row["matchedToCore"]),
                "支付宝用户ID": safe_str(row.get("alipayUserId")) or "",
                "登录手机": safe_str(row.get("loginPhone")) or "",
                "绑定手机": safe_str(row.get("boundPhone")) or "",
                "注册时间": safe_str(row.get("registrationTime")) or "",
                "注销时间": safe_str(row.get("cancellationTime")) or "",
                "可用余额(元)": _float_value(row.get("availableBalanceYuan")),
                "绑定银行卡数量": _int_value(row.get("bankCardCount")),
                "绑定银行卡": safe_str(row.get("bankCards")) or "",
                "关联账户": safe_str(row.get("linkedAccounts")) or "",
                "备注": safe_str(row.get("remarks")) or "",
                "来源文件": safe_str(row.get("sourceFile")) or "",
            }
        )
    return _frame(
        rows,
        [
            "序号",
            "主体姓名",
            "主体证件号",
            "是否命中主链",
            "支付宝用户ID",
            "登录手机",
            "绑定手机",
            "注册时间",
            "注销时间",
            "可用余额(元)",
            "绑定银行卡数量",
            "绑定银行卡",
            "关联账户",
            "备注",
            "来源文件",
        ],
    )


def _build_alipay_transaction_df(
    wallet_data: Dict[str, Any],
    artifact_details: Dict[str, List[Dict[str, Any]]],
) -> pd.DataFrame:
    rows = []
    enriched_rows = sorted(
        _enrich_subject_rows(wallet_data, artifact_details.get("alipayTransactionRows", [])),
        key=lambda item: (
            safe_str(item.get("subjectName")) or "",
            safe_str(item.get("createdAt")) or "",
            safe_str(item.get("transactionId")) or "",
        ),
    )
    for index, row in enumerate(enriched_rows, start=1):
        rows.append(
            {
                "序号": index,
                "主体姓名": row["subjectName"],
                "主体证件号": row["subjectId"],
                "是否命中主链": _yes_no(row["matchedToCore"]),
                "交易创建时间": safe_str(row.get("createdAt")) or "",
                "付款时间": safe_str(row.get("paidAt")) or "",
                "最近修改时间": safe_str(row.get("modifiedAt")) or "",
                "金额(元)": _float_value(row.get("amountYuan")),
                "收支方向": safe_str(row.get("direction")) or "",
                "是否有效交易": _yes_no(row.get("isEffective")),
                "交易状态": safe_str(row.get("status")) or "",
                "交易对方": safe_str(row.get("counterpartyName")) or "",
                "交易来源地": safe_str(row.get("transactionSource")) or "",
                "交易类型": safe_str(row.get("transactionType")) or "",
                "消费名称": safe_str(row.get("itemName")) or "",
                "支付宝交易号": safe_str(row.get("transactionId")) or "",
                "商户订单号": safe_str(row.get("merchantOrderNo")) or "",
                "支付方式": safe_str(row.get("paymentMethod")) or "",
                "清算流水号": safe_str(row.get("clearingSerialNo")) or "",
                "来源Sheet": safe_str(row.get("sourceSheet")) or "",
                "来源文件": safe_str(row.get("sourceFile")) or "",
            }
        )
    return _frame(
        rows,
        [
            "序号",
            "主体姓名",
            "主体证件号",
            "是否命中主链",
            "交易创建时间",
            "付款时间",
            "最近修改时间",
            "金额(元)",
            "收支方向",
            "是否有效交易",
            "交易状态",
            "交易对方",
            "交易来源地",
            "交易类型",
            "消费名称",
            "支付宝交易号",
            "商户订单号",
            "支付方式",
            "清算流水号",
            "来源Sheet",
            "来源文件",
        ],
    )


def _build_wechat_registration_df(
    wallet_data: Dict[str, Any],
    artifact_details: Dict[str, List[Dict[str, Any]]],
) -> pd.DataFrame:
    rows = []
    for index, row in enumerate(_enrich_subject_rows(wallet_data, artifact_details.get("wechatRegistrationRows", [])), start=1):
        rows.append(
            {
                "序号": index,
                "主体姓名": row["subjectName"],
                "主体证件号": row["subjectId"],
                "是否命中主链": _yes_no(row["matchedToCore"]),
                "匹配状态": _match_status_text(row.get("matchStatus")),
                "匹配依据": safe_str(row.get("matchBasis")) or "",
                "目录手机号": safe_str(row.get("phoneDirectory")) or "",
                "当前绑定手机号": safe_str(row.get("currentBoundPhone")) or "",
                "微信号": safe_str(row.get("wechatId")) or "",
                "微信别名": safe_str(row.get("wechatAlias")) or "",
                "昵称": safe_str(row.get("nickname")) or "",
                "注册时间": safe_str(row.get("registeredAt")) or "",
                "来源文件": safe_str(row.get("sourceFile")) or "",
            }
        )
    return _frame(
        rows,
        [
            "序号",
            "主体姓名",
            "主体证件号",
            "是否命中主链",
            "匹配状态",
            "匹配依据",
            "目录手机号",
            "当前绑定手机号",
            "微信号",
            "微信别名",
            "昵称",
            "注册时间",
            "来源文件",
        ],
    )


def _build_wechat_login_df(
    wallet_data: Dict[str, Any],
    artifact_details: Dict[str, List[Dict[str, Any]]],
) -> pd.DataFrame:
    rows = []
    enriched_rows = sorted(
        _enrich_subject_rows(wallet_data, artifact_details.get("wechatLoginRows", [])),
        key=lambda item: (
            safe_str(item.get("subjectName")) or "",
            safe_str(item.get("loginTime")) or "",
        ),
    )
    for index, row in enumerate(enriched_rows, start=1):
        rows.append(
            {
                "序号": index,
                "主体姓名": row["subjectName"],
                "主体证件号": row["subjectId"],
                "是否命中主链": _yes_no(row["matchedToCore"]),
                "匹配状态": _match_status_text(row.get("matchStatus")),
                "目录手机号": safe_str(row.get("phoneDirectory")) or "",
                "登录时间": safe_str(row.get("loginTime")) or "",
                "IP地址": safe_str(row.get("ipAddress")) or "",
                "来源文件": safe_str(row.get("sourceFile")) or "",
            }
        )
    return _frame(
        rows,
        [
            "序号",
            "主体姓名",
            "主体证件号",
            "是否命中主链",
            "匹配状态",
            "目录手机号",
            "登录时间",
            "IP地址",
            "来源文件",
        ],
    )


def _build_tenpay_registration_df(
    wallet_data: Dict[str, Any],
    artifact_details: Dict[str, List[Dict[str, Any]]],
) -> pd.DataFrame:
    rows = []
    for index, row in enumerate(_enrich_subject_rows(wallet_data, artifact_details.get("tenpayRegistrationRows", [])), start=1):
        rows.append(
            {
                "序号": index,
                "主体姓名": row["subjectName"],
                "主体证件号": row["subjectId"],
                "是否命中主链": _yes_no(row["matchedToCore"]),
                "财付通账号": safe_str(row.get("tenpayAccountAlias")) or "",
                "账户状态": safe_str(row.get("accountStatus")) or "",
                "注册时间": safe_str(row.get("registeredAt")) or "",
                "绑定手机": safe_str(row.get("boundPhone")) or "",
                "绑定状态": safe_str(row.get("bindingStatus")) or "",
                "开户行信息": safe_str(row.get("bankInfo")) or "",
                "银行卡数量": _int_value(row.get("bankCardCount")),
                "银行卡": safe_str(row.get("bankCards")) or "",
                "来源文件": safe_str(row.get("sourceFile")) or "",
            }
        )
    return _frame(
        rows,
        [
            "序号",
            "主体姓名",
            "主体证件号",
            "是否命中主链",
            "财付通账号",
            "账户状态",
            "注册时间",
            "绑定手机",
            "绑定状态",
            "开户行信息",
            "银行卡数量",
            "银行卡",
            "来源文件",
        ],
    )


def _build_tenpay_transaction_df(
    wallet_data: Dict[str, Any],
    artifact_details: Dict[str, List[Dict[str, Any]]],
) -> pd.DataFrame:
    rows = []
    enriched_rows = sorted(
        _enrich_subject_rows(wallet_data, artifact_details.get("tenpayTransactionRows", [])),
        key=lambda item: (
            safe_str(item.get("subjectName")) or "",
            safe_str(item.get("transactionTime")) or "",
            safe_str(item.get("transactionId")) or "",
        ),
    )
    for index, row in enumerate(enriched_rows, start=1):
        rows.append(
            {
                "序号": index,
                "主体姓名": row["subjectName"],
                "主体证件号": row["subjectId"],
                "是否命中主链": _yes_no(row["matchedToCore"]),
                "财付通账号": safe_str(row.get("tenpayAccountAlias")) or "",
                "交易时间": safe_str(row.get("transactionTime")) or "",
                "金额(元)": _float_value(row.get("amountYuan")),
                "借贷类型": safe_str(row.get("direction")) or "",
                "交易业务类型": safe_str(row.get("businessType")) or "",
                "交易用途类型": safe_str(row.get("purposeType")) or "",
                "对手方名称": safe_str(row.get("counterpartyName")) or "",
                "对手方ID": safe_str(row.get("counterpartyId")) or "",
                "对手方银行名称": safe_str(row.get("counterpartyBankName")) or "",
                "对手方银行卡号": safe_str(row.get("counterpartyBankCard")) or "",
                "对手方接收时间": safe_str(row.get("counterpartyReceivedAt")) or "",
                "对手方接收金额(元)": _float_value(row.get("counterpartyReceivedAmountYuan")),
                "交易单号": safe_str(row.get("transactionId")) or "",
                "大单号": safe_str(row.get("masterOrderNo")) or "",
                "用户银行卡号": safe_str(row.get("userBankCard")) or "",
                "渠道": safe_str(row.get("networkChannel")) or "",
                "备注1": safe_str(row.get("remark1")) or "",
                "备注2": safe_str(row.get("remark2")) or "",
                "来源文件": safe_str(row.get("sourceFile")) or "",
            }
        )
    return _frame(
        rows,
        [
            "序号",
            "主体姓名",
            "主体证件号",
            "是否命中主链",
            "财付通账号",
            "交易时间",
            "金额(元)",
            "借贷类型",
            "交易业务类型",
            "交易用途类型",
            "对手方名称",
            "对手方ID",
            "对手方银行名称",
            "对手方银行卡号",
            "对手方接收时间",
            "对手方接收金额(元)",
            "交易单号",
            "大单号",
            "用户银行卡号",
            "渠道",
            "备注1",
            "备注2",
            "来源文件",
        ],
    )


def _build_alert_df(wallet_data: Dict[str, Any]) -> pd.DataFrame:
    rows = []
    for index, alert in enumerate(wallet_data.get("alerts", []) or [], start=1):
        rows.append(
            {
                "序号": index,
                "风险级别": safe_str(alert.get("risk_level")) or "",
                "预警类型": safe_str(alert.get("alert_type")) or "",
                "主体": safe_str(alert.get("person")) or "",
                "对手方": safe_str(alert.get("counterparty")) or "",
                "日期": safe_str(alert.get("date")) or "",
                "金额(元)": _float_value(alert.get("amount")),
                "描述": safe_str(alert.get("description")) or "",
                "风险原因": safe_str(alert.get("risk_reason")) or "",
            }
        )
    return _frame(
        rows,
        ["序号", "风险级别", "预警类型", "主体", "对手方", "日期", "金额(元)", "描述", "风险原因"],
    )


def _build_unmatched_df(wallet_data: Dict[str, Any]) -> pd.DataFrame:
    rows = []
    for index, item in enumerate(wallet_data.get("unmatchedWechatAccounts", []) or [], start=1):
        rows.append(
            {
                "序号": index,
                "手机号": safe_str(item.get("phone")) or "",
                "微信号": safe_str(item.get("wxid")) or "",
                "别名": safe_str(item.get("alias")) or "",
                "昵称": safe_str(item.get("nickname")) or "",
                "注册时间": safe_str(item.get("registeredAt")) or "",
                "最新登录时间": safe_str(item.get("latestLoginAt")) or "",
                "登录次数": _int_value(item.get("loginEventCount")),
            }
        )
    return _frame(
        rows,
        ["序号", "手机号", "微信号", "别名", "昵称", "注册时间", "最新登录时间", "登录次数"],
    )


def _enrich_subject_rows(
    wallet_data: Dict[str, Any],
    rows: Iterable[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    subjects_by_id = wallet_data.get("subjectsById", {}) or {}
    enriched_rows: List[Dict[str, Any]] = []
    for row in rows or []:
        row_copy = dict(row)
        subject_id = safe_str(row_copy.get("subjectId")) or ""
        subject = subjects_by_id.get(subject_id) if subject_id else None
        row_copy["subjectId"] = subject_id
        row_copy["subjectName"] = (
            safe_str(subject.get("subjectName")) if isinstance(subject, dict) else None
        ) or safe_str(row_copy.get("subjectName")) or ""
        row_copy["matchedToCore"] = bool(subject.get("matchedToCore")) if isinstance(subject, dict) else False
        enriched_rows.append(row_copy)
    return enriched_rows


def _frame(rows: List[Dict[str, Any]], columns: List[str]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows, columns=columns)


def _yes_no(value: Any) -> str:
    return "是" if bool(value) else "否"


def _match_status_text(value: Any) -> str:
    normalized = (safe_str(value) or "").lower()
    if normalized == "matched":
        return "已归并"
    if normalized == "unmatched":
        return "未归并"
    return safe_str(value) or ""


def _int_value(value: Any, default: int = 0) -> int:
    converted = safe_int(value)
    return converted if converted is not None else default


def _float_value(value: Any, default: float = 0.0) -> float:
    converted = safe_float(value)
    return converted if converted is not None else default


def _format_worksheet(worksheet, column_count: int) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for row in worksheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _WRAP_ALIGNMENT

    for column_index in range(1, column_count + 1):
        column_letter = worksheet.cell(row=1, column=column_index).column_letter
        max_length = 0
        for cell in worksheet[column_letter]:
            text = "" if cell.value is None else str(cell.value)
            if len(text) > max_length:
                max_length = len(text)
            cell.alignment = _WRAP_ALIGNMENT
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)
