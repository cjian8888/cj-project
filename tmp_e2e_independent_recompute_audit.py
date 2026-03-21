#!/usr/bin/env python3
from __future__ import annotations

import json
import sqlite3
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DATA_ROOT = ROOT / "data"
OUTPUT_ROOT = ROOT / "output"
REPORT_JSON = OUTPUT_ROOT / "analysis_results" / "qa" / "report_package.json"
REPORT_OUT_TXT = (
    OUTPUT_ROOT / "analysis_results" / "qa" / "e2e_independent_recompute_audit_report.txt"
)
REPORT_OUT_JSON = (
    OUTPUT_ROOT / "analysis_results" / "qa" / "e2e_independent_recompute_audit_report.json"
)

BANK_COLUMNS = [
    "名称",
    "借贷标志",
    "交易金额",
    "交易时间",
    "交易对方名称",
    "交易对方账号",
    "交易对方卡号",
    "交易摘要",
    "交易流水号",
    "本方账号",
    "本方卡号",
    "查询卡号",
]

INCOME_HINTS = ("进", "贷", "收入", "入", "收")
EXPENSE_HINTS = ("出", "借", "支出", "付", "汇出", "转出")
COMPANY_MARKERS = (
    "公司",
    "集团",
    "研究所",
    "研究院",
    "中心",
    "银行",
    "医院",
    "大学",
    "学院",
    "协会",
)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().replace("\u200b", "").replace("\ufeff", "")
    return "" if text.lower() in {"nan", "none", "<na>"} else text


def clean_account(value: Any) -> str:
    text = clean_text(value).replace("\t", "").replace(" ", "")
    if text.endswith(".0"):
        text = text[:-2]
    return text


def money_to_cents(value: Any) -> int:
    text = clean_text(value).replace(",", "")
    if not text:
        return 0
    amount = Decimal(text).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return int(amount * 100)


def cents_to_money(cents: int) -> str:
    return f"{(Decimal(cents) / Decimal('100')).quantize(Decimal('0.01'))}"


def normalize_direction(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return "unknown"
    for token in INCOME_HINTS:
        if token in text:
            return "income"
    for token in EXPENSE_HINTS:
        if token in text:
            return "expense"
    return "unknown"


def is_company_name(name: str) -> bool:
    return any(marker in name for marker in COMPANY_MARKERS)


def iter_bank_records() -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    dedupe_keys: set[str] = set()
    for path in sorted(
        p for p in DATA_ROOT.rglob("*.xlsx") if "银行业金融机构交易流水" in str(p)
    ):
        workbook = load_workbook(path, read_only=False, data_only=True)
        sheet = workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
        header_row_index = None
        header_index: dict[str, int] = {}
        for idx, row in enumerate(rows):
            headers = [clean_text(value) for value in row]
            current_index = {name: pos for pos, name in enumerate(headers) if name}
            if "名称" in current_index and "交易金额" in current_index:
                header_row_index = idx
                header_index = current_index
                break
        if "名称" not in header_index or "交易金额" not in header_index:
            workbook.close()
            continue
        for excel_row_num, values in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
            entity_name = clean_text(values[header_index.get("名称", -1)])
            direction = normalize_direction(values[header_index.get("借贷标志", -1)])
            amount_cents = money_to_cents(values[header_index.get("交易金额", -1)])
            tx_time = clean_text(values[header_index.get("交易时间", -1)])
            counterparty_name = clean_text(values[header_index.get("交易对方名称", -1)])
            counterparty_account = clean_account(
                values[header_index.get("交易对方账号", -1)]
                or values[header_index.get("交易对方卡号", -1)]
            )
            own_account = clean_account(
                values[header_index.get("本方账号", -1)]
                or values[header_index.get("本方卡号", -1)]
                or values[header_index.get("查询卡号", -1)]
            )
            summary = clean_text(values[header_index.get("交易摘要", -1)])
            txid = clean_text(values[header_index.get("交易流水号", -1)])
            natural_key = "|".join(
                [
                    entity_name,
                    direction,
                    str(amount_cents),
                    tx_time,
                    counterparty_name,
                    summary,
                    own_account,
                ]
            )
            dedupe_key = (
                f"{path.name}|TXID|{txid}" if txid else f"{path.name}|NAT|{natural_key}"
            )
            if dedupe_key in dedupe_keys:
                continue
            dedupe_keys.add(dedupe_key)
            records.append(
                {
                    "source_file": path.name,
                    "entity_name": entity_name,
                    "direction": direction,
                    "amount_cents": amount_cents,
                    "transaction_time": tx_time,
                    "counterparty_name": counterparty_name,
                    "counterparty_account": counterparty_account,
                    "own_account": own_account,
                    "summary": summary,
                    "txid": txid,
                    "excel_row_num": excel_row_num,
                }
            )
        workbook.close()
    return records


def build_sqlite(records: list[dict[str, Any]]) -> sqlite3.Connection:
    conn = sqlite3.connect(":memory:")
    conn.execute(
        """
        create table bank_rows (
            entity_name text,
            direction text,
            amount_cents integer,
            transaction_time text,
            counterparty_name text,
            counterparty_account text,
            own_account text,
            summary text,
            txid text,
            source_file text,
            excel_row_num integer
        )
        """
    )
    conn.executemany(
        """
        insert into bank_rows (
            entity_name, direction, amount_cents, transaction_time, counterparty_name,
            counterparty_account, own_account, summary, txid, source_file, excel_row_num
        ) values (
            :entity_name, :direction, :amount_cents, :transaction_time, :counterparty_name,
            :counterparty_account, :own_account, :summary, :txid, :source_file, :excel_row_num
        )
        """,
        records,
    )
    return conn


def audit() -> dict[str, Any]:
    report = json.loads(REPORT_JSON.read_text(encoding="utf-8"))
    records = iter_bank_records()
    conn = build_sqlite(records)
    problems: list[dict[str, Any]] = []

    global_row = conn.execute(
        """
        select
            count(*) as transaction_count,
            sum(case when direction='income' then amount_cents else 0 end) as total_income_cents,
            sum(case when direction='expense' then amount_cents else 0 end) as total_expense_cents,
            count(distinct case when entity_name<>'' then entity_name end) as subject_count
        from bank_rows
        """
    ).fetchone()

    persons_count = conn.execute(
        "select count(distinct entity_name) from bank_rows where entity_name<>''"
    ).fetchone()[0]
    companies_count = conn.execute(
        "select entity_name from bank_rows where entity_name<>'' group by entity_name"
    ).fetchall()
    company_count = sum(1 for (name,) in companies_count if is_company_name(name))
    person_count = persons_count - company_count

    global_checks = [
        (
            "bank_transaction_count",
            int(report["coverage"]["bank_transaction_count"]),
            int(global_row[0] or 0),
        ),
        (
            "persons_count",
            int(report["coverage"]["persons_count"]),
            int(person_count),
        ),
        (
            "companies_count",
            int(report["coverage"]["companies_count"]),
            int(company_count),
        ),
        (
            "total_income_cents",
            int(
                sum(
                    money_to_cents(item.get("total_income"))
                    for item in report.get("person_dossiers", []) + report.get("company_dossiers", [])
                )
            ),
            int(global_row[1] or 0),
        ),
        (
            "total_expense_cents",
            int(
                sum(
                    money_to_cents(item.get("total_expense"))
                    for item in report.get("person_dossiers", []) + report.get("company_dossiers", [])
                )
            ),
            int(global_row[2] or 0),
        ),
    ]
    for metric, report_value, recomputed_value in global_checks:
        if report_value != recomputed_value:
            problems.append(
                {
                    "kind": "global_metric_mismatch",
                    "metric": metric,
                    "report_value": report_value,
                    "recomputed_value": recomputed_value,
                }
            )

    entity_mismatches: list[dict[str, Any]] = []
    for item in report.get("person_dossiers", []) + report.get("company_dossiers", []):
        entity_name = str(item.get("entity_name") or "").strip()
        row = conn.execute(
            """
            select
                count(*) as transaction_count,
                sum(case when direction='income' then amount_cents else 0 end) as total_income_cents,
                sum(case when direction='expense' then amount_cents else 0 end) as total_expense_cents
            from bank_rows
            where entity_name=?
            """,
            (entity_name,),
        ).fetchone()
        tx_count = int(row[0] or 0)
        income_cents = int(row[1] or 0)
        expense_cents = int(row[2] or 0)
        report_tx_count = int(item.get("transaction_count") or 0)
        report_income_cents = money_to_cents(item.get("total_income"))
        report_expense_cents = money_to_cents(item.get("total_expense"))
        if (
            tx_count != report_tx_count
            or income_cents != report_income_cents
            or expense_cents != report_expense_cents
        ):
            entity_mismatches.append(
                {
                    "entity_name": entity_name,
                    "report_transaction_count": report_tx_count,
                    "recomputed_transaction_count": tx_count,
                    "report_total_income": cents_to_money(report_income_cents),
                    "recomputed_total_income": cents_to_money(income_cents),
                    "report_total_expense": cents_to_money(report_expense_cents),
                    "recomputed_total_expense": cents_to_money(expense_cents),
                }
            )
    if entity_mismatches:
        problems.append(
            {
                "kind": "entity_rollup_mismatch",
                "count": len(entity_mismatches),
                "details": entity_mismatches[:20],
            }
        )

    top_turnover = conn.execute(
        """
        select
            entity_name,
            count(*) as transaction_count,
            sum(case when direction='income' then amount_cents else 0 end) as total_income_cents,
            sum(case when direction='expense' then amount_cents else 0 end) as total_expense_cents,
            sum(amount_cents) as turnover_cents
        from bank_rows
        where entity_name<>''
        group by entity_name
        order by turnover_cents desc
        limit 5
        """
    ).fetchall()

    conn.close()
    return {
        "checked_entity_count": len(report.get("person_dossiers", []))
        + len(report.get("company_dossiers", [])),
        "global_transaction_count": int(global_row[0] or 0),
        "global_total_income": cents_to_money(int(global_row[1] or 0)),
        "global_total_expense": cents_to_money(int(global_row[2] or 0)),
        "top_turnover_entities": [
            {
                "entity_name": row[0],
                "transaction_count": int(row[1] or 0),
                "total_income": cents_to_money(int(row[2] or 0)),
                "total_expense": cents_to_money(int(row[3] or 0)),
            }
            for row in top_turnover
        ],
        "problem_count": len(problems),
        "problems": problems,
        "verdict": (
            "异实现复算与正式语义层一致。"
            if not problems
            else "异实现复算发现与正式语义层不一致的结果。"
        ),
    }


def build_text_report(payload: dict[str, Any]) -> str:
    lines = [
        "======================================================================",
        "异实现复算审计报告",
        "======================================================================",
        f"独立读取原始银行流水后复算：总笔数={payload['global_transaction_count']}，总流入={payload['global_total_income']}，总流出={payload['global_total_expense']}。",
        f"共核对 {payload['checked_entity_count']} 个卷宗对象，发现问题 {payload['problem_count']} 项。",
        "",
        "Top 5 总流水对象:",
    ]
    for item in payload["top_turnover_entities"]:
        lines.append(
            f"- {item['entity_name']} | 笔数={item['transaction_count']} | 流入={item['total_income']} | 流出={item['total_expense']}"
        )
    if payload["problems"]:
        lines.append("")
        lines.append("问题明细:")
        for item in payload["problems"]:
            lines.append(f"- {item['kind']}: {json.dumps(item, ensure_ascii=False)}")
    lines.append("")
    lines.append(payload["verdict"])
    return "\n".join(lines)


def main() -> None:
    payload = audit()
    REPORT_OUT_JSON.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    text = build_text_report(payload)
    REPORT_OUT_TXT.write_text(text, encoding="utf-8")
    print(text)


if __name__ == "__main__":
    main()
