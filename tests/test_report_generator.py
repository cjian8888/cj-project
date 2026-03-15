#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""report_generator 老链路聚合 explainability 回归测试。"""

import os
import sys

import pandas as pd
import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from report_generator import (
    _generate_aggregation_summary_sheet,
    _generate_html_conclusion,
    _generate_report_conclusion,
    generate_excel_workbook,
    generate_word_report,
)


def _make_profiles():
    return {
        "张三": {
            "has_data": True,
            "summary": {
                "transaction_count": 12,
                "total_income": 300000,
                "total_expense": 200000,
            },
        }
    }


def _make_suspicions():
    return {
        "direct_transfers": [],
        "hidden_assets": {},
    }


def _make_derived_data():
    return {
        "aggregation": {
            "summary": {
                "极高风险实体数": 1,
                "高风险实体数": 0,
                "高优先线索实体数": 1,
            },
            "rankedEntities": [
                {
                    "name": "张三",
                    "riskScore": 88,
                    "riskConfidence": 0.91,
                    "riskLevel": "critical",
                    "highPriorityClueCount": 3,
                    "topEvidenceScore": 84,
                    "summary": "存在闭环和第三方中转",
                    "aggregationExplainability": {
                        "top_clues": [
                            {"description": "资金闭环: 张三 → 外围账户B → 张三"}
                        ]
                    },
                }
            ],
        }
    }


def test_generate_report_conclusion_prefers_aggregation_highlights():
    lines = _generate_report_conclusion(
        _make_profiles(),
        _make_suspicions(),
        ["张三"],
        [],
        derived_data=_make_derived_data(),
    )

    text = "\n".join(lines)
    assert "【聚合排序】" in text
    assert "张三(88.0分/置信度0.91)" in text
    assert "重点线索" in text


def test_generate_html_conclusion_prefers_aggregation_highlights():
    html = _generate_html_conclusion(
        _make_profiles(),
        _make_suspicions(),
        ["张三"],
        [],
        derived_data=_make_derived_data(),
    )

    assert "【聚合排序】" in html
    assert "张三(88.0分/置信度0.91)" in html
    assert "聚合排序识别出重点核查对象" in html


def test_generate_aggregation_summary_sheet_creates_excel_sheet(tmp_path):
    output_path = tmp_path / "aggregation.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _generate_aggregation_summary_sheet(writer, _make_derived_data())

    workbook = load_workbook(output_path)
    assert "聚合风险排序" in workbook.sheetnames
    assert "聚合风险重点对象" in workbook.sheetnames
    sheet = workbook["聚合风险排序"]
    values = [cell for row in sheet.iter_rows(values_only=True) for cell in row if cell is not None]
    assert "极高风险实体数" in values
    detail_values = [
        cell
        for row in workbook["聚合风险重点对象"].iter_rows(values_only=True)
        for cell in row
        if cell is not None
    ]
    assert "对象名称" in detail_values
    assert "张三" in detail_values


def test_generate_excel_workbook_hardens_aliases_and_quality(tmp_path):
    output_path = tmp_path / "底稿.xlsx"

    profiles = {
        "张三": {
            "has_data": True,
            "summary": {
                "transaction_count": 10,
                "total_income": 200000,
                "total_expense": 100000,
                "salary_ratio": 0.5,
            },
            "income_structure": {"salary_income": 100000},
            "fund_flow": {
                "third_party_amount": 1000,
                "third_party_income": 1234.56,
                "third_party_expense": 200.0,
                "third_party_income_count": 1,
                "third_party_expense_count": 1,
                "third_party_income_transactions": [
                    {
                        "日期": "2026-01-02T10:11:12",
                        "金额": 1234.56,
                        "摘要": "nan",
                        "对手方": "支付宝（中国）网络技术有限公司",
                        "source_file": "cleaned_data/个人/张三_合并流水.xlsx",
                        "source_row_index": 8,
                    }
                ],
                "third_party_expense_transactions": [
                    {
                        "日期": "2026-01-03T11:12:13",
                        "金额": 200.0,
                        "摘要": "None",
                        "对手方": "财付通",
                        "source_file": "cleaned_data/个人/张三_合并流水.xlsx",
                        "source_row_index": 9,
                    }
                ],
            },
            "wealth_management": {
                "wealth_purchase_transactions": [
                    {
                        "日期": "2026-01-04T09:10:11",
                        "金额": 50000.0,
                        "摘要": "理财购买",
                        "对手方": "银行理财账户",
                        "判断依据": "支出+基金",
                    }
                ],
                "wealth_redemption_transactions": [
                    {
                        "日期": "2026-01-05T09:10:11",
                        "金额": 10000.0,
                        "摘要": "nan",
                        "对手方": "银行理财账户",
                        "判断依据": "隐蔽赎回特征",
                    }
                ],
                "wealth_purchase": 50000.0,
                "wealth_purchase_count": 1,
                "wealth_redemption": 10000.0,
                "wealth_redemption_count": 1,
                "net_wealth_flow": 40000.0,
                "estimated_holding": 40000.0,
                "total_transactions": 2,
            },
        }
    }

    suspicions = {
        "cashCollisions": [
            {
                "person1": "张三",
                "person2": "李四",
                "time1": "2026-01-02T09:00:00",
                "time2": "2026-01-02T15:00:00",
                "amount1": 5000,
                "amount2": 5000,
                "timeDiff": 6,
                "riskLevel": "high",
                "riskReason": "现金进出高度吻合",
                "withdrawalSource": "张三流水.xlsx",
                "depositSource": "李四流水.xlsx",
            }
        ],
        "directTransfers": [
            {
                "from": "张三",
                "to": "某公司",
                "amount": 30000,
                "date": "2026-01-03",
                "direction": "out",
                "description": "往来款",
                "riskLevel": "high",
                "sourceFile": "张三流水.xlsx",
                "sourceRowIndex": 88,
            }
        ],
    }

    family_tree = {
        "张三": [
            {
                "姓名": "李四",
                "身份证号": "310101199001010011",
                "与户主关系": "配偶",
                "性别": "女",
                "出生日期": "1990-01-01",
                "户籍地": "上海",
                "数据来源": "户籍信息",
            }
        ]
    }

    family_assets = {
        "张三": {
            "家族成员": ["李四"],
            "房产套数": 1,
            "房产总价值": 120.0,
            "车辆数量": 1,
            "房产": [
                {
                    "产权人": "张三",
                    "房地坐落": "浦东新区示例路1号",
                    "建筑面积": "89.00平方米",
                    "交易金额": 0,
                    "规划用途": "居住",
                    "房屋性质": "商品房",
                    "登记时间": "2020-01-01",
                    "共有情况": "共同共有",
                    "共有人名称": "李四",
                    "权属状态": "现势",
                    "数据质量": "正常",
                }
            ],
            "车辆": [
                {
                    "所有人": "张三",
                    "号牌号码": "沪A12345",
                    "中文品牌": "特斯拉",
                    "车身颜色": "白色",
                    "初次登记日期": "2024-01-01",
                    "机动车状态": "正常",
                    "是否抵押质押": 1,
                    "能源种类": "纯电动",
                    "住所地址": "浦东新区示例路1号",
                }
            ],
        }
    }

    penetration_results = {
        "summary": {
            "个人→公司笔数": 1,
            "个人→公司总金额": 30000.0,
            "公司→个人笔数": 0,
            "公司→个人总金额": 0,
            "核心人员间笔数": 1,
            "核心人员间总金额": 1000.0,
            "涉案公司间笔数": 0,
            "涉案公司间总金额": 0,
        },
        "person_to_person": [
            {
                "发起方": "张三",
                "接收方": "李四",
                "日期": "2026-01-01T08:00:00",
                "收入": 1000.0,
                "支出": 0.0,
                "摘要": "nan",
                "交易对方原文": "李四",
            }
        ],
        "fund_cycles": [
            {
                "path": "张三 → 李四 → 张三",
                "length": 3,
                "risk_level": "critical",
                "risk_score": 88,
                "confidence": 0.9,
            }
        ],
        "pass_through_channels": [
            {
                "node": "中转账户",
                "node_type": "person",
                "inflow": 100000.0,
                "outflow": 95000.0,
                "ratio": 0.95,
                "net_retention": 5000.0,
                "risk_level": "high",
                "risk_score": 70,
                "confidence": 0.8,
                "evidence": ["进出比达到95%", "净沉淀较低"],
            }
        ],
        "hub_nodes": [
            {
                "node": "枢纽甲",
                "node_type": "person",
                "in_degree": 5,
                "out_degree": 6,
                "total_degree": 11,
            }
        ],
    }

    loan_results = {
        "bidirectional_flows": [
            {
                "person": "张三",
                "counterparty": "李四",
                "income_count": 2,
                "income_total": 20000.0,
                "expense_count": 1,
                "expense_total": 10000.0,
                "ratio": 0.5,
                "loan_type": "疑似借贷",
                "risk_level": "medium",
                "source_file": "cleaned_data/个人/张三_合并流水.xlsx",
                "source_row_index": 12,
            }
        ],
        "online_loan_platforms": [
            {
                "person": "张三",
                "platform": "消费贷",
                "counterparty": "平台A",
                "date": "2026-01-06",
                "amount": 3000.0,
                "direction": "income",
                "description": "授信入账",
                "risk_level": "medium",
                "source_file": "cleaned_data/个人/张三_合并流水.xlsx",
                "source_row_index": 20,
            }
        ],
    }

    income_results = {
        "regular_non_salary": [
            {
                "person": "张三",
                "counterparty": "代收付",
                "occurrences": 3,
                "avg_amount": 12000.0,
                "total_amount": 36000.0,
                "date_range": ["2025-01-01", "2025-03-01"],
                "possible_type": "个人大额转入（需关注）",
                "risk_level": "high",
                "source_file": "cleaned_data/个人/张三_合并流水.xlsx",
                "source_row_index": 30,
            }
        ],
        "large_individual_income": [
            {
                "person": "张三",
                "from_individual": "王五",
                "date": "2025-04-01",
                "amount": 50000.0,
                "description": "大额转入",
                "risk_level": "high",
                "source_file": "cleaned_data/个人/张三_合并流水.xlsx",
                "source_row_index": 31,
            }
        ],
        "unknown_source_income": [
            {
                "person": "张三",
                "counterparty": "(无)",
                "date": "2025-05-01",
                "amount": 0.0,
                "description": "强制扣划",
                "reason": "对手方信息缺失",
                "risk_level": "high",
                "source_file": "cleaned_data/个人/张三_合并流水.xlsx",
                "source_row_index": 32,
            }
        ],
        "same_source_multi": [
            {
                "person": "张三",
                "counterparty": "随行付",
                "count": 5,
                "total": 88000.0,
                "avg_amount": 17600.0,
                "date_range": ["2025-06-01", "2025-07-01"],
                "possible_type": "同源多次收入",
                "risk_level": "medium",
                "source_file": "cleaned_data/个人/张三_合并流水.xlsx",
                "source_row_index": 33,
            }
        ],
        "large_single_income": [
            {
                "person": "张三",
                "counterparty": "赵六",
                "date": "2025-08-01",
                "amount": 80000.0,
                "description": "单笔大额",
                "possible_type": "个人大额转入",
                "risk_level": "high",
                "source_file": "cleaned_data/个人/张三_合并流水.xlsx",
                "source_row_index": 34,
            }
        ],
        "potential_bribe_installment": [
            {
                "person": "张三",
                "counterparty": "商户A",
                "avg_monthly": 5000.0,
                "cv": 0.1,
                "months": 6,
                "count": 6,
                "total_amount": 30000.0,
                "risk_factors": "持续稳定入账",
                "risk_level": "high",
                "source_file": "cleaned_data/个人/张三_合并流水.xlsx",
                "source_row_index": 35,
            }
        ],
    }

    time_series_results = {
        "sudden_changes": [
            {
                "person": "张三",
                "date": "2025-09-01",
                "amount": 200000.0,
                "z_score": 12.5,
                "avg_before": 1000.0,
                "change_type": "income_spike",
                "risk_level": "high",
                "source_file": "cleaned_data/个人/张三_合并流水.xlsx",
                "source_row_index": 40,
            }
        ],
        "delayed_transfers": [
            {
                "person": "张三",
                "income_counterparty": "王五",
                "expense_counterparty": "李四",
                "delay_days": 2,
                "occurrences": 3,
                "avg_amount": 2000.0,
                "total_amount": 6000.0,
                "risk_level": "medium",
                "first_income_date": "2025-09-01T12:00:00",
                "source_file": "cleaned_data/个人/张三_合并流水.xlsx",
                "source_row_index": 41,
            }
        ],
    }

    derived_data = {
        "aggregation": _make_derived_data()["aggregation"],
        "correlation": {
            "summary": {
                "航班同行人数": 1,
                "铁路同行人数": 1,
                "同住宿人数": 1,
                "快递联系人数": 1,
            },
            "travel_companions": {
                "flight_companions": [
                    {
                        "person": "张三",
                        "companion_name": "李四",
                        "_travel_dt": "2025-11-26T00:00:00",
                        "travel_type": "航班",
                        "flight_no": "MU1234",
                        "source_file": "张三航班.xlsx",
                    }
                ],
                "rail_companions": [
                    {
                        "person": "张三",
                        "companion_name": "王五",
                        "travel_date": 20251127,
                        "travel_type": "火车",
                        "train_no": "G123",
                        "source_file": "张三铁路.xlsx",
                    }
                ],
                "fund_correlations": [
                    {
                        "person": "张三",
                        "companion": "李四",
                        "travel_date": 20251126,
                        "travel_type": "航班",
                        "transaction_date": "2025-11-20T09:00:00",
                        "days_diff": -6,
                        "timing": "先付款后同行",
                        "amount": 5000.0,
                        "direction": "expense",
                        "counterparty_raw": "李四",
                        "description": "同行前转账",
                        "risk_level": "medium",
                    }
                ],
            },
            "hotel_cohabitants": {
                "cohabitants": [
                    {
                        "person": "张三",
                        "cohabitant": "李四",
                        "_stay_dt": "2025-10-01T00:00:00",
                        "hotel": "测试酒店",
                    }
                ],
                "fund_correlations": [
                    {
                        "person": "张三",
                        "cohabitant": "李四",
                        "_stay_dt": "2025-10-01T00:00:00",
                        "transaction_date": "2025-10-02T10:00:00",
                        "days_diff": 1,
                        "timing": "同行后收款",
                        "amount": 8000.0,
                        "direction": "income",
                        "counterparty_raw": "李四",
                        "description": "住宿关联转账",
                        "risk_level": "high",
                    }
                ],
            },
            "express_contacts": {
                "express_contacts": [
                    {"person": "张三", "contact": "李四", "type": "收件人"}
                ],
                "frequent_addresses": [
                    {"name": "上海市浦东新区示例路1号", "count": 2, "persons": ["张三", "李四"]}
                ],
                "fund_correlations": [
                    {
                        "person": "张三",
                        "contact": "李四",
                        "transaction_date": "2025-11-03T09:00:00",
                        "amount": 9000.0,
                        "direction": "income",
                        "description": "快递联系人转账",
                        "frequency": 2,
                        "risk_level": "medium",
                    }
                ],
            },
        },
        "family_tree": family_tree,
        "all_family_summaries": {
            "张三家庭": {
                "total_assets": {
                    "bank_balance": 100000.0,
                    "property_value": 200000.0,
                    "vehicle_value": 30000.0,
                    "wealth_balance": 50000.0,
                    "total": 380000.0,
                    "property_count": 1,
                    "vehicle_count": 1,
                },
                "member_transfers": {
                    "张三": {
                        "to_family": 3000.0,
                        "from_family": 1000.0,
                        "net": -2000.0,
                        "transfer_details": [
                            {"counterparty": "李四", "amount": 3000.0, "date": "2025-11-01"},
                            {"counterparty": "李四", "amount": 1000.0, "date": "2025-11-05"},
                        ],
                    }
                },
            }
        },
    }

    result = generate_excel_workbook(
        profiles=profiles,
        suspicions=suspicions,
        output_path=str(output_path),
        family_tree=family_tree,
        family_assets=family_assets,
        penetration_results=penetration_results,
        loan_results=loan_results,
        income_results=income_results,
        time_series_results=time_series_results,
        derived_data=derived_data,
    )

    assert result == str(output_path)
    workbook = load_workbook(output_path)

    assert "家族关系图谱" in workbook.sheetnames
    assert "关联分析-汇总" in workbook.sheetnames
    assert "穿透-过账通道" in workbook.sheetnames
    assert "时序-资金突变" in workbook.sheetnames
    assert "异常收入-来源不明" in workbook.sheetnames
    assert "家庭总资产汇总" in workbook.sheetnames
    assert "成员间转账明细" in workbook.sheetnames

    direct_sheet = workbook["直接转账关系"]
    assert direct_sheet.freeze_panes == "A2"
    assert direct_sheet.auto_filter.ref is not None
    assert direct_sheet["G2"].value == "高风险"
    assert direct_sheet["H2"].value == "张三流水.xlsx"

    hub_sheet = workbook["穿透-枢纽节点"]
    assert hub_sheet["A2"].value == "枢纽甲"
    assert hub_sheet["B2"].value == "人员"

    pass_sheet = workbook["穿透-过账通道"]
    assert pass_sheet["A2"].value == "中转账户"
    assert pass_sheet["G2"].value == "高风险"

    time_sheet = workbook["时序-资金突变"]
    headers = [cell.value for cell in time_sheet[1]]
    assert "历史均值(元)" in headers
    assert time_sheet["F2"].value == "收入突增"
    assert time_sheet["G2"].value == "高风险"

    family_asset_sheet = workbook["家庭总资产汇总"]
    assert family_asset_sheet["A2"].value == "张三家庭"
    assert family_asset_sheet["F2"].value == 380000

    transfer_sheet = workbook["成员间转账明细"]
    assert transfer_sheet["A2"].value == "张三家庭"
    assert transfer_sheet["I2"].value == "李四"

    vehicle_sheet = workbook["车辆明细"]
    assert vehicle_sheet["H2"].value == "是"

    family_tree_sheet = workbook["家族关系图谱"]
    assert family_tree_sheet["C2"].value == "310101199001010011"

    anomaly_sheet = workbook["异常收入-汇总"]
    anomaly_values = [
        cell
        for row in anomaly_sheet.iter_rows(values_only=True)
        for cell in row
        if isinstance(cell, str)
    ]
    assert "来源不明收入" in anomaly_values
    assert "nan" not in anomaly_values
    assert "None" not in anomaly_values

    third_party_sheet = workbook["第三方支付-收入"]
    third_party_headers = [cell.value for cell in third_party_sheet[1]]
    third_party_row = next(third_party_sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    third_party_map = dict(zip(third_party_headers, third_party_row))
    assert third_party_map["来源文件"] == "张三_合并流水.xlsx"
    third_party_values = [
        cell
        for row in third_party_sheet.iter_rows(values_only=True)
        for cell in row
        if isinstance(cell, str)
    ]
    assert "nan" not in third_party_values
    assert "None" not in third_party_values
    if "摘要" in third_party_headers:
        assert third_party_map["摘要"] in ("", None)


def test_generate_word_report_prefers_aggregation_highlights(tmp_path):
    docx = pytest.importorskip("docx")
    output_path = tmp_path / "aggregation.docx"

    result = generate_word_report(
        _make_profiles(),
        _make_suspicions(),
        ["张三"],
        [],
        output_path=str(output_path),
        family_summary={},
        family_assets={},
        cleaned_data={},
        derived_data=_make_derived_data(),
    )

    assert result == str(output_path)
    document = docx.Document(result)
    text = "\n".join(p.text for p in document.paragraphs if p.text.strip())
    assert "【聚合排序】" in text
    assert "张三(88.0分/置信度0.91)" in text
    assert "聚合排序识别出重点核查对象 张三(88.0分)" in text
